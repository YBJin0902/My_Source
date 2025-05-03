import string
from openpyxl import Workbook

def create_excel_from_matrix(matrix, filename):
    wb = Workbook()
    ws = wb.active

    # Label rows and columns with alphabets
    labels = list(string.ascii_uppercase)
    max_index = len(matrix)

    # Write column labels
    for i in range(max_index):
        ws.cell(row=1, column=i+2, value=labels[i])

    # Write row labels and matrix data
    for i in range(max_index):
        ws.cell(row=i+2, column=1, value=labels[i])  # Row labels
        for j in range(max_index):
            ws.cell(row=i+2, column=j+2, value=matrix[i][j])

    wb.save(filename)

# Example matrix
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

create_excel_from_matrix(matrix, "output.xlsx")